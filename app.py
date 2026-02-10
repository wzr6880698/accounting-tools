import os
import sys
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import shutil
from datetime import datetime, timedelta
import warnings
import re
import zipfile
import streamlit as st

# ============================
# 网页配置
# ============================
st.set_page_config(
    page_title="会计分录凭证生成系统",
    page_icon="🧾",
    layout="wide"
)

st.title("🧾 会计分录凭证生成系统")
st.markdown("---")

# ============================
# 侧边栏：文件上传区域
# ============================
st.sidebar.header("📁 步骤 1: 上传文件")

# 上传凭证文件
entries_file = st.sidebar.file_uploader(
    "1️⃣ 上传会计分录文件 (支持 .xls / .xlsx / XML)",
    type=['xlsx', 'xls', 'xlsm', 'xml'],
    key="entries_file"
)

# 上传收款收据模板
receipt_template = st.sidebar.file_uploader(
    "2️⃣ 上传收款收据模板",
    type=['xlsx', 'xls'],
    key="receipt_template"
)

# 上传领款凭证模板
payment_template = st.sidebar.file_uploader(
    "3️⃣ 上传领款凭证模板",
    type=['xlsx', 'xls'],
    key="payment_template"
)

st.sidebar.header("⚙️ 步骤 2: 生成凭证")
generate_button = st.sidebar.button("🚀 开始生成凭证", type="primary")

# ============================
# 以下是你的核心逻辑函数
# ============================


def parse_accounting_entries(file_path):
    """解析会计分录导出文件（终极三引擎+命名空间版）"""
    df = None
    engine_used = "未知"
    
    try:
        # 检查 lxml 是否可用
        try:
            import lxml
            lxml_version = lxml.__version__
            print(f"检测到 lxml 库，版本: {lxml_version}")
        except ImportError:
            print("❌ 严重错误：未安装 lxml 库！XML 文件将无法读取！")
            st.error("❌ 系统错误：云端环境缺少 lxml 库，请检查 requirements.txt。")

        # ==========================================
        # 第一步：用记事本方式读取文件头，判断真实格式
        # ==========================================
        file_header = ""
        try:
            with open(file_path, 'rb') as f:
                # 读取前200字节并转码，用来判断文件头
                file_header = f.read(200).decode('utf-8', errors='ignore') 
        except:
            # 读取失败，直接跳过检测，交给后续逻辑处理
            file_header = ""
        
        # print(f"调试: 检测文件头前100字符: {file_header[:100]}")

        # ==========================================
        # 第二步：根据格式选择读取方式
        # ==========================================
        
        # 情况 A：Excel 2003 XML 格式 (以 <?xml 或 <Workbook 开头)
        if file_header.startswith('<?xml') or file_header.startswith('<Workbook'):
            print("✅ 检测到 Excel 2003 XML 格式，尝试使用 read_xml 读取...")
            try:
                # Excel 2003 XML 带有特定的命名空间，必须指定 xpath 才能准确读取
                # 命名空间定义
                namespaces = {
                    'ss': 'urn:schemas-microsoft-com:office:spreadsheet'
                }
                
                # 使用 Pandas 读取 XML
                # xpath=".//ss:Row" 表示只读取 Row 标签
                # stylesheet=None (默认)
                df = pd.read_xml(
                    file_path, 
                    xpath=".//ss:Row",
                    namespaces=namespaces
                )
                
                engine_used = "read_xml (Excel 2003 XML)"
                print(f"✅ XML 读取成功 (带命名空间): {len(df)}行 x {len(df.columns)}列")
                
                # 特殊处理：XML 读出来的第一行通常是表头，需要转换
                # 如果列名是 0, 1, 2... 这种数字，说明第一行数据没被当作表头
                # 我们尝试将第一行设为列名
                if len(df.columns) >= 1 and str(df.columns[0]).isdigit():
                    print("检测到列名为数字，尝试提取第一行作为表头...")
                    # 保存第一行作为新表头
                    new_header = df.iloc[0].values
                    # 删除第一行
                    df = df[1:]
                    # 重命名
                    df.columns = new_header
                
            except Exception as e:
                print(f"❌ XML 读取 (带命名空间) 失败: {e}")
                # 如果带命名空间失败，尝试不带命名空间的（兼容性回退）
                print("尝试回退到不带命名空间的 read_xml...")
                try:
                    df = pd.read_xml(file_path)
                    engine_used = "read_xml (通用模式)"
                    print(f"✅ 通用模式读取成功: {len(df)}行")
                except Exception as e2:
                    print(f"❌ 通用模式也失败: {e2}")
                    raise Exception("XML 格式读取彻底失败。请检查文件是否损坏或 requirements.txt 中是否包含 lxml。")

        # 情况 B：标准的 .xls 文件
        elif file_path.endswith('.xls'):
            # 先试 xlrd (针对真正的 .xls)
            try:
                df = pd.read_excel(file_path, dtype=str, engine='xlrd')
                engine_used = "xlrd"
            except Exception as e:
                print(f"xlrd 引擎读取失败: {e}")
                print("尝试切换回 openpyxl 引擎 (可能是 .xlsx 格式伪装成了 .xls)...")
                # 如果 xlrd 失败，可能是伪装的 .xls，试 openpyxl
                df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
                engine_used = "openpyxl (回退)"

        # 情况 C：标准的 .xlsx 文件
        else:
            df = pd.read_excel(file_path, dtype=str, engine='openpyxl')
            engine_used = "openpyxl"

        print(f"成功读取Excel文件: {os.path.basename(file_path)} (引擎: {engine_used})")
        print(f"数据维度: {len(df)}行 × {len(df.columns)}列")

        # ... 以下代码保持不变：列名匹配与数据处理 ...
        # 显示前几行数据以了解结构
        print("\n前3行数据预览:")
        print(df.head(3))
        print("\n列名:")
        for i, col in enumerate(df.columns):
            print(f"  列{i + 1}: {col}")

        # 重命名列，假设文件有6列
        if len(df.columns) >= 6:
            # 使用前6列
            df = df.iloc[:, :6]
            df.columns = ['日期', '凭证字号', '摘要', '科目', '借方金额', '贷方金额']
            print("已使用前6列作为标准列名")
        else:
            print(f"文件只有{len(df.columns)}列，尝试匹配列名...")

            # 尝试匹配列名
            column_mapping = {}
            for i, col in enumerate(df.columns):
                col_str = str(col).lower()
                if any(keyword in col_str for keyword in ['日期', 'date']):
                    column_mapping[i] = '日期'
                    print(f"  列{i + 1}匹配为: 日期")
                elif any(keyword in col_str for keyword in ['凭证', 'voucher', '字号', '凭证号']):
                    column_mapping[i] = '凭证字号'
                    print(f"  列{i + 1}匹配为: 凭证字号")
                elif any(keyword in col_str for keyword in ['摘要', 'summary', 'remark', '内容']):
                    column_mapping[i] = '摘要'
                    print(f"  列{i + 1}匹配为: 摘要")
                elif any(keyword in col_str for keyword in ['科目', 'account', 'subject', '科目名称', '科目代码']):
                    column_mapping[i] = '科目'
                    print(f"  列{i + 1}匹配为: 科目")
                elif any(keyword in col_str for keyword in ['借方', 'debit', '借方金额']):
                    column_mapping[i] = '借方金额'
                    print(f"  列{i + 1}匹配为: 借方金额")
                elif any(keyword in col_str for keyword in ['贷方', 'credit', '贷方金额']):
                    column_mapping[i] = '贷方金额'
                    print(f"  列{i + 1}匹配为: 贷方金额")
                else:
                    # 如果没有匹配，使用默认顺序
                    default_names = ['日期', '凭证字号', '摘要', '科目', '借方金额', '贷方金额']
                    if i < len(default_names):
                        column_mapping[i] = default_names[i]
                        print(f"  列{i + 1}默认设置为: {default_names[i]}")

            # 创建新的DataFrame
            new_data = {}
            for i, new_name in column_mapping.items():
                if i < len(df.columns):
                    new_data[new_name] = df.iloc[:, i]

            df = pd.DataFrame(new_data)

        # 填充合并单元格
        df = fill_merged_cells(df)

        # 清理数据
        for col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace(['nan', 'NaN', 'None', 'null', '', 'NaT'], pd.NA)

        # 将金额列转换为数值类型
        if '借方金额' in df.columns:
            df['借方金额'] = pd.to_numeric(df['借方金额'].astype(str).str.replace(',', ''), errors='coerce')
        if '贷方金额' in df.columns:
            df['贷方金额'] = pd.to_numeric(df['贷方金额'].astype(str).str.replace(',', ''), errors='coerce')

        # 填充NaN为0
        if '借方金额' in df.columns:
            df['借方金额'] = df['借方金额'].fillna(0)
        if '贷方金额' in df.columns:
            df['贷方金额'] = df['贷方金额'].fillna(0)

        # 只保留有金额的行
        if '借方金额' in df.columns and '贷方金额' in df.columns:
            initial_count = len(df)
            df = df[(df['借方金额'] != 0) | (df['贷方金额'] != 0)]
            filtered_count = len(df)
            print(f"金额清理: 从{initial_count}行过滤到{filtered_count}行有效数据")

        print(f"\n最终数据维度: {len(df)}行 × {len(df.columns)}列")

        # 检查是否有库存现金科目
        cash_rows = df[df['科目'].astype(str).str.contains('1001|库存现金', na=False)]
        print(f"找到 {len(cash_rows)} 行包含库存现金科目的记录")

        # 显示一些包含库存现金的样本数据，用于调试
        if len(cash_rows) > 0:
            print("\n库存现金科目样本数据:")
            for i, (idx, row) in enumerate(cash_rows.head(5).iterrows()):
                print(
                    f"  样本{i + 1}: 日期={row['日期']}, 科目={row['科目']}, 借方={row['借方金额']}, 贷方={row['贷方金额']}")

        return df

    except Exception as e:
        print(f"读取Excel文件彻底失败: {e}")
        import traceback
        traceback.print_exc()
        # 在网页上也显示详细错误，方便调试
        st.error(f"❌ 读取文件失败: {str(e)}")
        
        # 给出更具体的提示
        if "lxml" in str(e):
            st.error("💡 错误提示：系统缺少 lxml 库。请确保 requirements.txt 中包含 lxml，并重新部署。")
        
        return None


def fill_merged_cells(df):
    """填充合并单元格：向前填充日期和凭证字号列"""
    # 检查必要的列是否存在
    if '日期' in df.columns:
        # 向前填充日期列
        initial_nulls = df['日期'].isna().sum()
        df['日期'] = df['日期'].ffill()
        filled_nulls = df['日期'].isna().sum()
        if initial_nulls > filled_nulls:
            print(f"日期列: 填充了 {initial_nulls - filled_nulls} 个空值")

    if '凭证字号' in df.columns:
        # 向前填充凭证字号列
        initial_nulls = df['凭证字号'].isna().sum()
        df['凭证字号'] = df['凭证字号'].ffill()
        filled_nulls = df['凭证字号'].isna().sum()
        if initial_nulls > filled_nulls:
            print(f"凭证字号列: 填充了 {initial_nulls - filled_nulls} 个空值")

    return df


def get_grouped_entries(df):
    """将会计分录按凭证分组，找出完整的借贷关系"""
    if df is None or len(df) == 0:
        return {}

    # 按日期和凭证字号分组
    grouped = {}

    for idx, row in df.iterrows():
        date_val = row.get('日期', '')
        voucher_val = row.get('凭证字号', '')

        # 清理日期和凭证字号
        if pd.isna(date_val):
            date_val = ''
        if pd.isna(voucher_val):
            voucher_val = f"未命名_{idx}"

        key = f"{date_val}_{voucher_val}"

        if key not in grouped:
            grouped[key] = {
                '日期': date_val,
                '凭证字号': voucher_val,
                '摘要': row.get('摘要', ''),
                'entries': []  # 存储该凭证下的所有分录
            }

        grouped[key]['entries'].append({
            '科目': row.get('科目', ''),
            '摘要': row.get('摘要', ''),
            '借方金额': row.get('借方金额', 0),
            '贷方金额': row.get('贷方金额', 0)
        })

    print(f"已将会计分录分组为 {len(grouped)} 个凭证")
    return grouped


def extract_counterparty_for_cash_debit(group_entries):
    """从现金借方对应的贷方科目中提取二级科目（个人姓名）"""
    # 查找现金借方分录
    cash_debit_entry = None
    for entry in group_entries:
        account = str(entry.get('科目', ''))
        debit_amount = entry.get('借方金额', 0)

        if ('1001' in account or '库存现金' in account) and debit_amount > 0:
            cash_debit_entry = entry
            break

    if not cash_debit_entry:
        return "未知交款人"

    # 查找对应的贷方分录
    for entry in group_entries:
        credit_amount = entry.get('贷方金额', 0)
        account = str(entry.get('科目', ''))

        # 跳过现金科目本身
        if '1001' in account or '库存现金' in account:
            continue

        if credit_amount > 0:
            # 尝试从科目中提取二级科目（个人姓名）
            account_str = str(entry.get('科目', ''))

            # 尝试提取"-"后面的部分
            if '-' in account_str:
                parts = account_str.split('-')
                if len(parts) > 1:
                    name_part = parts[-1].strip()
                    # 去除可能的科目代码
                    name_part = re.sub(r'\d+\s*', '', name_part)
                    if name_part and len(name_part) <= 10:  # 假设名字不会太长
                        return name_part

            # 尝试提取"/"后面的部分
            if '/' in account_str:
                parts = account_str.split('/')
                if len(parts) > 1:
                    name_part = parts[-1].strip()
                    # 去除可能的科目代码
                    name_part = re.sub(r'\d+\s*', '', name_part)
                    if name_part and len(name_part) <= 10:
                        return name_part

            # 尝试提取" "空格后面的部分
            if ' ' in account_str:
                parts = account_str.split(' ')
                for part in parts:
                    # 跳过纯数字部分（科目代码）
                    if not re.match(r'^\d+$', part):
                        # 检查是否包含中文字符
                        if re.search(r'[\u4e00-\u9fff]', part):
                            return part[:10]  # 限制长度

    # 如果没有提取到，从摘要中提取
    summary = cash_debit_entry.get('摘要', '')
    return extract_counterparty_from_summary(summary)


def extract_counterparty_for_cash_credit(group_entries):
    """从现金贷方对应的借方科目中提取二级科目（个人姓名）"""
    # 查找现金贷方分录
    cash_credit_entry = None
    for entry in group_entries:
        account = str(entry.get('科目', ''))
        credit_amount = entry.get('贷方金额', 0)

        if ('1001' in account or '库存现金' in account) and credit_amount > 0:
            cash_credit_entry = entry
            break

    if not cash_credit_entry:
        return "未知领款人"

    # 查找对应的借方分录
    for entry in group_entries:
        debit_amount = entry.get('借方金额', 0)
        account = str(entry.get('科目', ''))

        # 跳过现金科目本身
        if '1001' in account or '库存现金' in account:
            continue

        if debit_amount > 0:
            # 尝试从科目中提取二级科目（个人姓名）
            account_str = str(entry.get('科目', ''))

            # 尝试提取"-"后面的部分
            if '-' in account_str:
                parts = account_str.split('-')
                if len(parts) > 1:
                    name_part = parts[-1].strip()
                    # 去除可能的科目代码
                    name_part = re.sub(r'\d+\s*', '', name_part)
                    if name_part and len(name_part) <= 10:
                        return name_part

            # 尝试提取"/"后面的部分
            if '/' in account_str:
                parts = account_str.split('/')
                if len(parts) > 1:
                    name_part = parts[-1].strip()
                    # 去除可能的科目代码
                    name_part = re.sub(r'\d+\s*', '', name_part)
                    if name_part and len(name_part) <= 10:
                        return name_part

    # 如果没有提取到，从摘要中提取
    summary = cash_credit_entry.get('摘要', '')
    return extract_counterparty_from_summary(summary)


def extract_counterparty_from_summary(summary):
    """从摘要中提取对方单位名称"""
    if not summary or pd.isna(summary):
        return "未知"

    summary = str(summary)

    # 常见关键词
    keywords = ["向", "从", "支付", "付", "收", "收到", "借", "还款", "付款", "给", "交", "还"]

    for keyword in keywords:
        if keyword in summary:
            # 提取关键词后面的部分
            parts = summary.split(keyword, 1)
            if len(parts) > 1:
                counterparty = parts[1].strip()
                # 去除常见的尾随词
                end_words = ["借款", "款项", "费用", "款", "现金", "金额", "租金", "运费", "包装费", "电费", "社保",
                             "费", "利息"]
                for end_word in end_words:
                    if counterparty.endswith(end_word):
                        counterparty = counterparty[:-len(end_word)].strip()

                if counterparty:
                    return counterparty[:15]  # 限制长度

    # 如果没有匹配到，返回原始摘要（截断）
    if len(summary) <= 15:
        return summary
    else:
        return summary[:12] + "..."


def get_business_date(base_date_str, is_receipt=True):
    """根据基础日期获取业务日期"""
    try:
        # 将字符串日期转换为datetime对象
        if isinstance(base_date_str, str) and base_date_str and base_date_str.strip():
            try:
                # 尝试多种日期格式
                date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%Y.%m.%d', '%Y%m%d']
                base_date = None

                for fmt in date_formats:
                    try:
                        # 截取前10个字符尝试解析
                        date_str = base_date_str.strip()[:10]
                        base_date = datetime.strptime(date_str, fmt)
                        break
                    except:
                        continue

                if not base_date:
                    # 如果无法解析，使用当前日期
                    base_date = datetime.now()
            except:
                base_date = datetime.now()
        else:
            base_date = datetime.now()

        year = base_date.year
        month = base_date.month

        if is_receipt:
            # 收款日期：当月1日（如果是周末，顺延到下一个工作日）
            day = 1
        else:
            # 领款日期：当月15日（如果是周末，顺延到下一个工作日）
            day = 15

        # 创建日期
        try:
            business_date = datetime(year, month, day)
        except ValueError:
            # 如果日期无效（如2月30日），使用当月最后一天
            import calendar
            last_day = calendar.monthrange(year, month)[1]
            day = min(day, last_day)
            business_date = datetime(year, month, day)

        # 调整周末：周六(5)周日(6)
        while business_date.weekday() >= 5:  # 5=周六, 6=周日
            business_date += timedelta(days=1)

        return business_date

    except Exception as e:
        print(f"获取业务日期失败: {e}")
        # 返回当前日期作为备选
        return datetime.now()


def convert_to_chinese_amount(num):
    """将数字金额转换为中文大写金额"""
    try:
        # 定义数字对应的大写
        chinese_digits = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
        chinese_units = ['', '拾', '佰', '仟']
        chinese_big_units = ['', '万', '亿']

        # 处理负数
        if num < 0:
            return "负" + convert_to_chinese_amount(abs(num))

        # 处理零
        if num == 0:
            return "零元整"

        # 分离整数和小数部分
        integer_part = int(num)
        decimal_part = round((num - integer_part) * 100)

        # 处理整数部分
        if integer_part == 0:
            chinese_integer = "零"
        else:
            chinese_integer = ""
            str_num = str(integer_part)

            # 分组处理（每4位一组）
            groups = []
            while str_num:
                groups.append(str_num[-4:])
                str_num = str_num[:-4]
            groups.reverse()

            for i, group in enumerate(groups):
                group_chinese = ""
                group_len = len(group)
                zero_in_group = False

                for j, digit in enumerate(group):
                    digit_int = int(digit)
                    unit_pos = group_len - j - 1

                    if digit_int != 0:
                        # 如果前面有零，先添加零
                        if zero_in_group:
                            group_chinese += '零'
                            zero_in_group = False

                        # 添加数字（十位上的"壹"通常省略）
                        if not (digit_int == 1 and unit_pos == 1 and j == 0):
                            group_chinese += chinese_digits[digit_int]

                        # 添加单位（除了个位）
                        if unit_pos > 0:
                            group_chinese += chinese_units[unit_pos]
                    else:
                        zero_in_group = True

                # 如果组不为空，添加大单位
                if group_chinese or (i == len(groups) - 1 and chinese_integer == ""):
                    chinese_integer += group_chinese
                    if i < len(groups) - 1:
                        chinese_integer += chinese_big_units[len(groups) - i - 1]

        # 如果整数部分为空，则添加"零"
        if not chinese_integer:
            chinese_integer = "零"

        # 处理小数部分
        chinese_decimal = ""
        if decimal_part > 0:
            jiao = decimal_part // 10
            fen = decimal_part % 10

            if jiao > 0:
                chinese_decimal += chinese_digits[jiao] + "角"
            if fen > 0:
                chinese_decimal += chinese_digits[fen] + "分"

        # 组合结果
        if chinese_decimal:
            result = chinese_integer + "元" + chinese_decimal
        else:
            result = chinese_integer + "元整"

        # 清理可能的"零零"
        while "零零" in result:
            result = result.replace("零零", "零")

        # 清理可能的"零元"
        if result.startswith("零元"):
            result = result[1:]  # 去掉开头的零

        # 清理可能的"零万"、"零亿"
        result = result.replace("零万", "万").replace("零亿", "亿")

        # 清理末尾的零
        if result.endswith("零"):
            result = result.rstrip("零")

        return result

    except Exception as e:
        print(f"转换金额大写失败: {e}")
        return f"（金额转换错误: {num}）"


def format_date_cell(cell, date_value):
    """设置日期单元格格式为'某年某月某日'并居中"""
    if isinstance(date_value, datetime):
        cell.value = date_value
        # 设置日期格式为中文年月日
        cell.number_format = 'yyyy年mm月dd日'
    else:
        cell.value = date_value

    # 设置居中对齐
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # 可以添加一些样式
    cell.font = Font(name='宋体', size=11)


def format_amount_cell(cell, amount):
    """设置金额单元格格式"""
    cell.value = amount
    # 设置数字格式为会计格式
    cell.number_format = '"¥"#,##0.00'
    # 设置右对齐
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.font = Font(name='宋体', size=11)


def format_text_cell(cell, text):
    """设置文本单元格格式"""
    if pd.isna(text):
        cell.value = ""
    else:
        cell.value = str(text)
    # 设置左对齐
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.font = Font(name='宋体', size=11)


def generate_receipt(template_path, output_path, entry_data, group_entries):
    """生成收款收据"""
    try:
        # 复制模板文件
        shutil.copyfile(template_path, output_path)

        # 加载工作簿
        wb = load_workbook(output_path)
        ws = wb.active

        # 获取业务日期（当月1日，如果是周末则顺延）
        business_date = get_business_date(entry_data.get('日期'), is_receipt=True)

        # 填写数据
        # B2: 日期
        format_date_cell(ws['B2'], business_date)

        # B3: 交款单位（从现金借方对应的贷方科目二级科目中提取）
        counterparty = extract_counterparty_for_cash_debit(group_entries)
        format_text_cell(ws['B3'], counterparty)

        # B4: 交款项目（摘要）
        summary = entry_data.get('摘要', '')
        format_text_cell(ws['B4'], summary)

        # C5: 小写金额
        amount = entry_data.get('金额', 0)
        print(f"调试信息: 收款收据金额 = {amount}")
        format_amount_cell(ws['C5'], amount)

        # B5: 大写金额
        chinese_amount = convert_to_chinese_amount(amount)
        format_text_cell(ws['B5'], chinese_amount)

        # 保存文件
        wb.save(output_path)
        print(f"已生成收款收据: {os.path.basename(output_path)}")
        return True

    except Exception as e:
        print(f"生成收款收据失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def generate_payment_voucher(template_path, output_path, entry_data, group_entries):
    """生成领款凭证"""
    try:
        # 复制模板文件
        shutil.copyfile(template_path, output_path)

        # 加载工作簿
        wb = load_workbook(output_path)
        ws = wb.active

        # 获取业务日期（当月15日，如果是周末则顺延）
        business_date = get_business_date(entry_data.get('日期'), is_receipt=False)

        # 填写数据
        # B2: 日期
        format_date_cell(ws['B2'], business_date)

        # B3: 领款单位（从现金贷方对应的借方科目二级科目中提取）
        counterparty = extract_counterparty_for_cash_credit(group_entries)
        format_text_cell(ws['B3'], counterparty)

        # B4: 领款原因（摘要）
        summary = entry_data.get('摘要', '')
        format_text_cell(ws['B4'], summary)

        # C5: 小写金额
        # 修正：使用'金额'而不是'金額'
        amount = entry_data.get('金额', 0)
        print(f"调试信息: 领款凭证金额 = {amount}")
        format_amount_cell(ws['C5'], amount)

        # B5: 大写金额
        chinese_amount = convert_to_chinese_amount(amount)
        format_text_cell(ws['B5'], chinese_amount)

        # 保存文件
        wb.save(output_path)
        print(f"已生成领款凭证: {os.path.basename(output_path)}")
        return True

    except Exception as e:
        print(f"生成领款凭证失败: {e}")
        import traceback
        traceback.print_exc()
        return False


def process_accounting_entries(entries_file, receipt_template, payment_template, output_dir):
    """处理会计分录，生成相应凭证"""
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)

    # 解析会计分录
    df = parse_accounting_entries(entries_file)

    if df is None or len(df) == 0:
        print("没有找到会计分录数据")
        return 0, 0, []  # 返回生成的文件列表

    print("\n开始处理会计分录...")
    print("-" * 80)

    # 按凭证分组
    grouped_data = get_grouped_entries(df)

    receipt_count = 0
    payment_count = 0
    generated_files = []  # 记录生成的文件

    # 遍历每个分组（每个凭证）
    for key, group_data in grouped_data.items():
        try:
            date = group_data.get('日期', '')
            voucher_no = group_data.get('凭证字号', '')
            group_entries = group_data.get('entries', [])

            # 检查该分组中是否有库存现金分录
            has_cash = False
            cash_entry = None
            cash_direction = None

            for entry in group_entries:
                account = str(entry.get('科目', ''))
                debit_amount = entry.get('借方金额', 0)
                credit_amount = entry.get('贷方金额', 0)

                if ('1001' in account or '库存现金' in account):
                    has_cash = True
                    if debit_amount > 0:
                        cash_entry = entry
                        cash_direction = '借方'
                        break
                    elif credit_amount > 0:
                        cash_entry = entry
                        cash_direction = '贷方'
                        break

            if not has_cash or not cash_entry:
                continue

            # 准备数据
            # 确定金额：如果现金在借方，金额为借方金额；如果在贷方，金额为贷方金额
            if cash_direction == '借方':
                amount = cash_entry.get('借方金额', 0)
            else:
                amount = cash_entry.get('贷方金额', 0)

            entry_data = {
                '日期': date,
                '凭证字号': voucher_no,
                '摘要': cash_entry.get('摘要', ''),
                '科目': cash_entry.get('科目', ''),
                '金额': amount,
                '方向': cash_direction
            }

            print(f"调试信息: 处理凭证 {voucher_no}, 日期 {date}, 方向 {cash_direction}, 金额 {amount}")

            # 生成文件名
            # 清理日期字符串
            if date and not pd.isna(date):
                date_str = re.sub(r'[^\d]', '', str(date))[:8]
            else:
                date_str = datetime.now().strftime('%Y%m%d')

            # 清理凭证字号
            if voucher_no and not pd.isna(voucher_no):
                voucher_no_clean = re.sub(r'[^\w\u4e00-\u9fff]', '', str(voucher_no))
                if not voucher_no_clean:
                    voucher_no_clean = f"凭证{receipt_count + payment_count + 1}"
            else:
                voucher_no_clean = f"凭证{receipt_count + payment_count + 1}"

            if cash_direction == '借方':
                # 生成收款收据
                output_filename = f"收款收据_{date_str}_{voucher_no_clean}.xlsx"
                output_path = os.path.join(output_dir, output_filename)

                # 检查文件是否已存在，避免覆盖
                counter = 1
                while os.path.exists(output_path):
                    output_filename = f"收款收据_{date_str}_{voucher_no_clean}_{counter}.xlsx"
                    output_path = os.path.join(output_dir, output_filename)
                    counter += 1

                if generate_receipt(receipt_template, output_path, entry_data, group_entries):
                    receipt_count += 1
                    generated_files.append(output_path)
                    print(
                        f"  ✓ {date} {voucher_no} - {entry_data['摘要'][:30]}... - 金额: ¥{entry_data['金额']:,.2f} (库存现金借方)")

            elif cash_direction == '贷方':
                # 生成领款凭证
                output_filename = f"领款凭证_{date_str}_{voucher_no_clean}.xlsx"
                output_path = os.path.join(output_dir, output_filename)

                # 检查文件是否已存在，避免覆盖
                counter = 1
                while os.path.exists(output_path):
                    output_filename = f"领款凭证_{date_str}_{voucher_no_clean}_{counter}.xlsx"
                    output_path = os.path.join(output_dir, output_filename)
                    counter += 1

                if generate_payment_voucher(payment_template, output_path, entry_data, group_entries):
                    payment_count += 1
                    generated_files.append(output_path)
                    print(
                        f"  ✓ {date} {voucher_no} - {entry_data['摘要'][:30]}... - 金额: ¥{entry_data['金额']:,.2f} (库存现金贷方)")

        except Exception as e:
            print(f"处理凭证分组 {key} 时出错: {e}")
            import traceback
            traceback.print_exc()
            continue

    print("-" * 80)
    print(f"处理完成！")
    print(f"成功生成 {receipt_count} 个收款收据")
    print(f"成功生成 {payment_count} 个领款凭证")
    print(f"所有文件已保存到: {output_dir}")

    # 如果没有生成任何文件，可能是列名不匹配
    if receipt_count == 0 and payment_count == 0:
        print("\n注意：未生成任何凭证，可能是数据格式问题。")
        print("请检查数据格式，确保包含以下列：")
        print("1. 日期")
        print("2. 凭证字号")
        print("3. 摘要")
        print("4. 科目（包含'1001'或'库存现金'）")
        print("5. 借方金额")
        print("6. 贷方金额")
        print("\n如果以上列名不匹配，请确保您的Excel文件有正确的列标题。")
    
    return receipt_count, payment_count, generated_files


# ============================
# 网页主逻辑
# ============================

# 显示使用说明
with st.expander("📖 使用说明", expanded=False):
    st.markdown("""
    ### 操作步骤：
    1. 在左侧侧边栏依次上传三个文件：
       - 会计分录文件（支持 .xls / .xlsx / XML 格式）
       - 收款收据模板（Excel）
       - 领款凭证模板（Excel）
    
    2. 点击「开始生成凭证」按钮
    
    3. 等待处理完成，系统会自动生成所有凭证文件
    
    4. 点击「下载所有凭证文件（ZIP）」将所有文件打包下载
    
    ### 文件格式要求：
    - 会计分录文件应包含以下列：日期、凭证字号、摘要、科目、借方金额、贷方金额
    - 科目列需要包含"1001"或"库存现金"才会生成对应凭证
    """)

# 当用户点击生成按钮时
if generate_button:
    # 检查是否上传了所有必要文件
    if not entries_file:
        st.error("❌ 请先上传会计分录文件！")
    elif not receipt_template:
        st.error("❌ 请先上传收款收据模板！")
    elif not payment_template:
        st.error("❌ 请先上传领款凭证模板！")
    else:
        # 开始处理
        st.info("🔄 正在处理中，请稍候...")
        
        # 创建临时目录
        temp_dir = os.path.join(os.getcwd(), "temp_output")
        os.makedirs(temp_dir, exist_ok=True)
        
        try:
            # 保存上传的文件到临时目录
            # ==========================================
            # 关键修复：保留原始文件的后缀名 (例如 .xls)
            # 这样程序才能判断是用 xlrd 还是 openpyxl 引擎
            # ==========================================
            file_ext = os.path.splitext(entries_file.name)[1]
            entries_path = os.path.join(temp_dir, f"entries{file_ext}")
            
            receipt_path = os.path.join(temp_dir, "receipt_template.xlsx")
            payment_path = os.path.join(temp_dir, "payment_template.xlsx")
            
            with open(entries_path, "wb") as f:
                f.write(entries_file.getbuffer())
            with open(receipt_path, "wb") as f:
                f.write(receipt_template.getbuffer())
            with open(payment_path, "wb") as f:
                f.write(payment_template.getbuffer())
            
            # 调用核心处理函数
            receipt_count, payment_count, generated_files = process_accounting_entries(
                entries_path, receipt_path, payment_path, temp_dir
            )
            
            # 显示处理结果
            st.success(f"✅ 处理完成！")
            
            col1, col2 = st.columns(2)
            col1.metric("收款收据", f"{receipt_count} 个")
            col2.metric("领款凭证", f"{payment_count} 个")
            
            # 显示生成的文件列表
            if generated_files:
                st.subheader("📋 生成的文件列表")
                for file_path in generated_files:
                    st.text(f"• {os.path.basename(file_path)}")
                
                # 创建 ZIP 压缩包供下载
                zip_filename = f"凭证文件_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                zip_path = os.path.join(temp_dir, zip_filename)
                
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file_path in generated_files:
                        zipf.write(file_path, os.path.basename(file_path))
                
                # 提供下载按钮
                with open(zip_path, 'rb') as f:
                    st.download_button(
                        label="📥 下载所有凭证文件（ZIP）",
                        data=f.read(),
                        file_name=zip_filename,
                        mime="application/zip"
                    )
            
            else:
                st.warning("⚠️ 未生成任何凭证，请检查数据格式！")
                st.info("""
                ### 可能的原因：
                1. 会计分录文件中没有包含"1001"或"库存现金"科目的记录
                2. 文件列名不匹配，请确保包含：日期、凭证字号、摘要、科目、借方金额、贷方金额
                3. 数据为空或格式不正确
                """)
        
        except Exception as e:
            st.error(f"❌ 处理过程中出错：{str(e)}")
            import traceback
            traceback.print_exc()

# 页脚
st.markdown("---")
st.markdown("<div style='text-align: center; color: gray;'>💡 提示：所有处理都在本地完成，数据不会上传到任何服务器</div>", unsafe_allow_html=True)
st.markdown("<div style='text-align: center; color: gray;'>⚠️ 重要：如遇到 .xls 文件问题，请先转换为 .xlsx 格式</div>", unsafe_allow_html=True)
